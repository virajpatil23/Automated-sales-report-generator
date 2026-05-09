# ============================================
# Automated Sales Report Generator
# Author: Viraj Raosaheb Patil
# Description: Reads sales CSV, generates KPI
# summary Excel report, and emails it
# ============================================

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import os

# STEP 1: Load and clean data

def load_and_clean(filepath):
    df = pd.read_csv(filepath)
    df['Date'] = pd.to_datetime(df['Date'])
    df.dropna(inplace=True)
    df = df[df['Revenue'] > 0]
    df = df[df['Quantity'] > 0]
    print(f"Loaded {len(df)} rows from {filepath}")
    return df

# STEP 2: Generate KPIs

def generate_kpis(df):
    kpis = {
        'Total Revenue':     round(df['Revenue'].sum(), 2),
        'Total Orders':      len(df),
        'Avg Order Value':   round(df['Revenue'].mean(), 2),
        'Top Product':       df.groupby('Product')['Revenue'].sum().idxmax(),
        'Top Region':        df.groupby('Region')['Revenue'].sum().idxmax(),
        'Report Generated':  datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    return kpis

# STEP 3: Build Excel report

def build_excel(df, kpis, output_path):
    wb = Workbook()

    # ── Sheet 1: KPI Summary ──
    ws1 = wb.active
    ws1.title = "KPI Summary"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font  = Font(bold=True, size=14, color="1F4E79")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    ws1['A1'] = "Sales Performance Report"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:B1')

    ws1['A3'] = "KPI"
    ws1['B3'] = "Value"
    for col in ['A', 'B']:
        ws1[f'{col}3'].font = header_font
        ws1[f'{col}3'].fill = header_fill
        ws1[f'{col}3'].alignment = Alignment(horizontal='center')

    for i, (key, val) in enumerate(kpis.items(), start=4):
        ws1[f'A{i}'] = key
        ws1[f'B{i}'] = str(val)
        ws1[f'A{i}'].border = border
        ws1[f'B{i}'].border = border

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 30

    # ── Sheet 2: Revenue by Product ──
    ws2 = wb.create_sheet("Revenue by Product")
    product_summary = df.groupby('Product')['Revenue'].sum().reset_index()
    product_summary.columns = ['Product', 'Total Revenue']
    product_summary = product_summary.sort_values('Total Revenue', ascending=False)

    ws2['A1'] = "Product"
    ws2['B1'] = "Total Revenue (₹)"
    for col in ['A', 'B']:
        ws2[f'{col}1'].font = header_font
        ws2[f'{col}1'].fill = header_fill
        ws2[f'{col}1'].alignment = Alignment(horizontal='center')

    for i, row in enumerate(product_summary.itertuples(), start=2):
        ws2[f'A{i}'] = row.Product
        ws2[f'B{i}'] = row._2
        ws2[f'A{i}'].border = border
        ws2[f'B{i}'].border = border

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 22

    # ── Sheet 3: Revenue by Region ──
    ws3 = wb.create_sheet("Revenue by Region")
    region_summary = df.groupby('Region')['Revenue'].sum().reset_index()
    region_summary.columns = ['Region', 'Total Revenue']
    region_summary = region_summary.sort_values('Total Revenue', ascending=False)

    ws3['A1'] = "Region"
    ws3['B1'] = "Total Revenue (₹)"
    for col in ['A', 'B']:
        ws3[f'{col}1'].font = header_font
        ws3[f'{col}1'].fill = header_fill
        ws3[f'{col}1'].alignment = Alignment(horizontal='center')

    for i, row in enumerate(region_summary.itertuples(), start=2):
        ws3[f'A{i}'] = row.Region
        ws3[f'B{i}'] = row._2
        ws3[f'A{i}'].border = border
        ws3[f'B{i}'].border = border

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 22

    # ── Sheet 4: Raw Data ──
    ws4 = wb.create_sheet("Raw Data")
    headers = list(df.columns)
    for j, h in enumerate(headers, start=1):
        cell = ws4.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            ws4.cell(row=i, column=j, value=val)

    for col in range(1, len(headers)+1):
        ws4.column_dimensions[get_column_letter(col)].width = 18

    wb.save(output_path)
    print(f"Excel report saved: {output_path}")

# MAIN

if __name__ == "__main__":
    # Config
    INPUT_CSV   = 'sales_data.csv'
    OUTPUT_FILE = f'sales_report_{datetime.now().strftime("%Y%m%d")}.xlsx'

    # Run pipeline
    df   = load_and_clean(INPUT_CSV)
    kpis = generate_kpis(df)
    build_excel(df, kpis, OUTPUT_FILE)

    print("\n── KPI Summary ──")
    for k, v in kpis.items():
        print(f"  {k}: {v}")

    # Cron schedule (run this script weekly):
    # 0 8 * * 1 /usr/bin/python3 /path/to/sales_report_generator.py